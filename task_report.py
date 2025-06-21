# -*- coding: utf-8 -*-
"""Generate a report of tasks that still need to be returned.

The script reads two Excel files:
 - `ZS-沪乍杭-线路任务单一览表-补定测.xlsx`: column A lists every task that must be returned. Each
   value in column A should contain a form number plus a task index,
   e.g. "0101" meaning form 01 task 01.
 - `对应表格.xlsx`: the first column is a file name, the second column
   contains returned task identifiers. The returned task identifier has
   the same four character format where the first two digits are the
   form number and the last two digits are the task index. If the last
   two characters are ``AL`` it means all tasks for that form have been
   returned. ``0000`` means the mapping is unknown and is ignored.

The script compares both files and prints the tasks that are still
unreturned.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from typing import Iterable, List
import os

try:
    import pandas as pd
    import matplotlib.pyplot as plt
except Exception:
    sys.stderr.write("pandas and matplotlib are required to run this script\n")
    raise


DEFAULT_ZS_FILE = "沪乍杭-线路任务单一览表-补定测.xlsx"
DEFAULT_RETURNED_FILE = "对应表格.xlsx"

# Match a task code consisting of two digits for the form number followed by two
# digits or ``AL`` for the task index. The negative look-behind/ahead ensure the
# code is not part of a longer sequence of digits.
TASK_CODE_RE = re.compile(r"(?<!\d)(\d{2})(\d{2}|AL)(?!\d)", re.IGNORECASE)


def _parse_task_code(value: object):
    """Parse a task code like ``0101`` or ``03AL``.

    Returns ``(form_no, index)`` or ``None`` if the value cannot be
    parsed or represents an unknown mapping (``0000``).
    """
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.upper() == "0000":
        return None
    m = TASK_CODE_RE.search(text)
    if not m:
        return None
    form_no, index = m.groups()
    index = index.upper()
    return form_no, index


TASK_TEXT_RE = re.compile(r"(\d+)[-－](\d+)")


def _load_required_tasks(filename: str = DEFAULT_ZS_FILE):
    """Load required tasks from the ZS workbook.

    Returns a mapping ``{(form_no, index): category}`` where ``category`` is
    either ``"\u9053\u8def\u6284\u5e73"`` (road levelling), ``"\u6838\u8865\u5730\u5f62"``
    (terrain check) or ``None`` when no keywords match.
    """
    try:
        xls = pd.ExcelFile(filename)
    except FileNotFoundError:
        sys.stderr.write(f"File '{filename}' not found\n")
        return {}
    except Exception as e:
        sys.stderr.write(f"Failed to read '{filename}': {e}\n")
        return {}

    ROAD_LEVEL = "\u9053\u8def\u6284\u5e73"  # 道路抄平
    TERRAIN = "\u6838\u8865\u5730\u5f62"  # 核补地形

    tasks: dict[tuple[str, str], str | None] = {}
    for sheet in xls.sheet_names:
        if not re.match(r"\d+", sheet):
            continue  # skip summary or other sheets
        df = xls.parse(sheet, header=None)
        for row in df.itertuples(index=False):
            cells = [c for c in row if isinstance(c, str)]
            if not cells:
                continue
            # Determine classification based on keywords in the row
            classification = None
            row_text = "".join(cells)
            if "\u6284\u5e73" in row_text or "\u8d85\u5e73" in row_text:
                classification = ROAD_LEVEL
            elif "\u6838\u8865" in row_text:
                classification = TERRAIN
            for cell in cells:
                m = TASK_TEXT_RE.search(cell)
                if m:
                    form_no, index = m.groups()
                    tasks[(form_no.zfill(2), index.zfill(2))] = classification
                    break
    return tasks


def _load_returned_tasks(filename: str = DEFAULT_RETURNED_FILE):
    try:
        df = pd.read_excel(filename, header=None)
    except FileNotFoundError:
        sys.stderr.write(f"File '{filename}' not found\n")
        return {}
    except Exception as e:
        sys.stderr.write(f"Failed to read '{filename}': {e}\n")
        return {}

    returned = defaultdict(set)  # form_no -> set of indices or {"AL"}
    for value in df.iloc[:, 1].dropna().tolist():
        text = str(value)
        for form_no, index in TASK_CODE_RE.findall(text.upper()):
            # ignore unknown mapping "0000"
            if form_no == "00" and index == "00":
                continue
            returned[form_no].add(index.upper())
    return returned


def compute_unreturned(required_tasks, returned_tasks):
    """Return a list of unreturned task identifiers."""
    remaining = []
    for (form_no, index) in sorted(required_tasks):
        returned_indices = returned_tasks.get(form_no)
        if not returned_indices:
            remaining.append((form_no, index))
            continue
        if "AL" in returned_indices:
            continue  # all tasks for this form already returned
        if index not in returned_indices:
            remaining.append((form_no, index))
    return remaining


def compute_summary(required_tasks, returned_tasks):
    """Compute remaining tasks and per-form summary.

    ``required_tasks`` is a mapping ``{(form_no, index): category}``.
    ``returned_tasks`` maps form numbers to sets of returned indices or
    ``{"AL"}`` when all are returned.

    Returns ``(remaining, summary, categories, detail)`` where ``categories``
    lists totals and returned counts for each recognised category and ``detail``
    contains one row per classified task with a returned flag.
    """
    required_map: defaultdict[str, dict[str, str | None]] = defaultdict(dict)
    for (form_no, index), category in required_tasks.items():
        required_map[form_no][index] = category

    summary = []
    remaining = []
    class_total: defaultdict[str, int] = defaultdict(int)
    class_returned: defaultdict[str, int] = defaultdict(int)
    detail_rows: List[dict] = []
    for form_no in sorted(required_map):
        tasks = required_map[form_no]
        required_indices = set(tasks.keys())
        returned_indices = returned_tasks.get(form_no, set())
        if "AL" in returned_indices:
            returned_count = len(required_indices)
            missing_indices: List[str] = []
        else:
            missing_indices = sorted(required_indices - returned_indices)
            returned_count = len(required_indices) - len(missing_indices)
        ratio = (
            returned_count / len(required_indices)
            if required_indices
            else 0
        )
        summary.append(
            {
                "Form": form_no,
                "Required": len(required_indices),
                "Returned": returned_count,
                "Missing": " ".join(missing_indices),
                "Ratio": ratio,
            }
        )
        for idx, cat in tasks.items():
            returned_flag = "AL" in returned_indices or idx in returned_indices
            if cat:
                class_total[cat] += 1
                if returned_flag:
                    class_returned[cat] += 1
                detail_rows.append(
                    {
                        "Category": cat,
                        "Task": f"{form_no}{idx}",
                        "Returned": "Yes" if returned_flag else "No",
                    }
                )
            if idx in missing_indices:
                remaining.append((form_no, idx))
    category_summary = [
        {"Category": k, "Total": class_total[k], "Returned": class_returned.get(k, 0)}
        for k in class_total
    ]
    return remaining, summary, category_summary, detail_rows


def save_report(
    remaining: Iterable[tuple[str, str]],
    summary: List[dict],
    categories: List[dict],
    details: List[dict],
    output_file: str,
) -> None:
    """Save a visual report of remaining tasks to ``output_file``.

    The report contains a sheet listing each remaining task, a summary
    sheet with counts per form number and return ratios, a category total
    sheet and a category detail sheet showing each classified task.
    """
    if not remaining:
        df = pd.DataFrame(columns=["Form", "TaskIndex"])
    else:
        df = pd.DataFrame(remaining, columns=["Form", "TaskIndex"])
    summary_df = pd.DataFrame(summary)
    category_df = pd.DataFrame(categories)
    detail_df = pd.DataFrame(details)
    try:
        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Remaining")
            summary_df.to_excel(writer, index=False, sheet_name="Summary")
            if not category_df.empty:
                category_df.to_excel(writer, index=False, sheet_name="Category")
            if not detail_df.empty:
                detail_df.to_excel(writer, index=False, sheet_name="CategoryDetail")
    except Exception as e:
        sys.stderr.write(f"Failed to write report '{output_file}': {e}\n")


def save_bar_chart(summary: List[dict], chart_file: str) -> None:
    """Save a bar chart of return ratios for each form."""
    if not summary:
        return
    forms = [item["Form"] for item in summary]
    ratios = [item["Ratio"] for item in summary]
    plt.figure(figsize=(max(6, len(forms) * 0.6), 4))
    plt.bar(forms, ratios, color="skyblue")
    plt.ylim(0, 1)
    plt.xlabel("Form")
    plt.ylabel("Returned / Required")
    plt.title("Return Ratio by Form")
    plt.tight_layout()
    try:
        plt.savefig(chart_file)
    except Exception as e:
        sys.stderr.write(f"Failed to save chart '{chart_file}': {e}\n")
    finally:
        plt.close()


def mark_zs_file(
    zs_file: str, returned_tasks: dict[str, set[str]]
) -> tuple[str, dict[str, dict[str, int]]]:
    """Write return flags into the ZS workbook and update the summary sheet.

    The function scans sheets named with digits, writes ``"是"`` (yes) or
    ``"否"`` (no) in the ``"是否返回"`` column for each task and fills the
    "汇总" sheet with the total number of tasks and returned counts.  It
    then reads columns B, D and E of that sheet to aggregate totals by
    task type. The processed workbook is saved with ``_processed``
    appended to the original file name and the path of the written file
    is returned together with the type summary.
    """
    try:
        import openpyxl
    except Exception:
        sys.stderr.write("openpyxl is required to mark the workbook\n")
        return ""

    wb = openpyxl.load_workbook(zs_file)
    form_stats: dict[str, list[int]] = {}

    for sheet_name in wb.sheetnames:
        if not re.match(r"\d+", sheet_name):
            continue
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        try:
            return_col = header.index("是否返回") + 1
        except ValueError:
            return_col = 2
        task_col = 1
        total = 0
        returned_count = 0
        for row in range(2, ws.max_row + 1):
            cells = [ws.cell(row=row, column=task_col).value]
            if cells[0] is None:
                # if column A empty, check rest of row
                cells = [c.value for c in ws[row]]
            code = None
            for cell in cells:
                if isinstance(cell, str):
                    m = TASK_TEXT_RE.search(cell)
                    if m:
                        code = (m.group(1).zfill(2), m.group(2).zfill(2))
                        break
            if not code:
                continue
            form_no, index = code
            returned_indices = returned_tasks.get(form_no, set())
            returned_flag = ("AL" in returned_indices) or (index in returned_indices)
            ws.cell(row=row, column=return_col, value="是" if returned_flag else "否")
            total += 1
            if returned_flag:
                returned_count += 1

        base_form = re.match(r"(\d+)", sheet_name).group(1).zfill(2)
        prev = form_stats.get(base_form, [0, 0])
        form_stats[base_form] = [prev[0] + total, prev[1] + returned_count]

    type_summary: dict[str, dict[str, int]] = {}
    if "汇总" in wb.sheetnames:
        ws = wb["汇总"]
        for row in range(2, ws.max_row + 1):
            form_val = ws.cell(row=row, column=1).value
            if form_val is None:
                continue
            form_str = str(form_val).strip()
            if not form_str or not re.match(r"\d+", form_str):
                continue
            form_no = re.match(r"\d+", form_str).group(0).zfill(2)
            stats = form_stats.get(form_no)
            if stats:
                ws.cell(row=row, column=4, value=stats[0])
                ws.cell(row=row, column=5, value=stats[1])

        for row in range(2, ws.max_row + 1):
            task_type = ws.cell(row=row, column=2).value
            total = ws.cell(row=row, column=4).value
            returned_val = ws.cell(row=row, column=5).value
            if task_type is None or total is None or returned_val is None:
                continue
            try:
                total_i = int(total)
                returned_i = int(returned_val)
            except Exception:
                continue


            task_key = str(task_type).strip()
            if "核补地形" in task_key and task_key != "核补地形":
                task_key = "核补地形"

            if total_i == 0 and returned_i == 0:
                # Skip categories with no tasks
                continue

            info = type_summary.setdefault(task_key, {"提出数量": 0, "符合要求数量": 0})

            info["提出数量"] += total_i
            info["符合要求数量"] += returned_i

    base, ext = os.path.splitext(zs_file)
    output_file = f"{base}_processed{ext}"
    wb.save(output_file)
    return output_file, type_summary


def main(argv=None):
    parser = argparse.ArgumentParser(description="Report unreturned tasks")
    parser.add_argument(
        "--zs",
        default=DEFAULT_ZS_FILE,
        help="Excel file listing required tasks (default: %(default)s)",
    )
    parser.add_argument(
        "--returned",
        default=DEFAULT_RETURNED_FILE,
        help="Excel file listing already returned tasks (default: %(default)s)",
    )
    parser.add_argument(
        "-o",
        "--output",
        help="Write an Excel report to this file",
    )
    parser.add_argument(
        "-c",
        "--chart",
        help="Save a bar chart of return ratios to this PNG file",
    )
    args = parser.parse_args(argv)

    required = _load_required_tasks(args.zs)
    returned = _load_returned_tasks(args.returned)
    if not required:
        sys.stderr.write("No required tasks loaded or file missing.\n")
    if not returned:
        sys.stderr.write("No returned task information loaded or file missing.\n")

    remaining, summary, categories, detail = compute_summary(required, returned)
    processed, type_summary = mark_zs_file(args.zs, returned)
    if processed:
        print(f"Processed workbook saved to {processed}")
        if type_summary:
            print("\u4efb\u52a1\u7c7b\u578b\u7edf\u8ba1:")  # 任务类型统计
            for t, info in type_summary.items():
                print(f"{t}: \u63d0\u51fa{info['提出数量']}\u6761, \u8fd4\u56de{info['符合要求数量']}\u6761")
    if args.output:
        save_report(remaining, summary, categories, detail, args.output)
        print(f"Report written to {args.output}")
    if args.chart:
        save_bar_chart(summary, args.chart)
        print(f"Chart saved to {args.chart}")

    for idx, item in enumerate(summary, 1):
        if not item["Missing"]:
            print(f"{idx}: \u5168\u9f50")  # 全齐
        else:
            print(f"{idx}: \u7f3a{item['Missing']}")  # 缺...

    if categories:
        print("\u5206\u7c7b\u7edf\u8ba1:")  # 分类统计
        for cat in categories:
            print(
                f"{cat['Category']}: {cat['Returned']}/{cat['Total']}"
            )


if __name__ == "__main__":
    main()
