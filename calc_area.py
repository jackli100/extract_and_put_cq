"""Simple script to sum the values in column F of 0615.xlsx.

The script first coerces all cells in column F to numeric format (without
decimals) using ``openpyxl``.  It then reads the file with ``pandas`` and
outputs the total sum of the column.
"""
import sys

try:
    import pandas as pd
    from openpyxl import load_workbook
except Exception as e:
    sys.stderr.write("Both pandas and openpyxl are required to run this script\n")
    raise


def main(filename="0615.xlsx", column_index=5):
    # Convert the F column to integer format using openpyxl before reading with
    # pandas.  This ensures values stored as text are consistently treated as
    # numbers and rounding errors are avoided.
    try:
        wb = load_workbook(filename)
        ws = wb.active
        col_letter = chr(ord('A') + column_index)
        for cell in ws[f"{col_letter}1":f"{col_letter}{ws.max_row}"]:
            c = cell[0]
            if c.value is None:
                continue
            try:
                num = int(float(str(c.value).strip()))
            except Exception:
                # Skip cells that cannot be converted
                continue
            c.value = num
            c.number_format = '0'
        wb.save(filename)
    except FileNotFoundError:
        sys.stderr.write(f"File '{filename}' not found\n")
        return 1
    except Exception as e:
        sys.stderr.write(f"Failed to prepare '{filename}': {e}\n")
        return 1

    try:
        df = pd.read_excel(filename)
    except FileNotFoundError:
        sys.stderr.write(f"File '{filename}' not found\n")
        return 1
    except Exception as e:
        sys.stderr.write(f"Failed to read '{filename}': {e}\n")
        return 1

    if column_index >= len(df.columns):
        sys.stderr.write(
            f"File only has {len(df.columns)} columns; cannot read column {column_index + 1}\n"
        )
        return 1

    # Read the target column and attempt to parse numeric values. If the cell
    # contains extra text such as units, we ignore it by coercing to NaN. This
    # avoids misinterpreting strings like "(m2)" as numbers.
    col = df.iloc[:, column_index].astype(str).str.replace(',', '').str.strip()
    mask = col.str.match(r'^[-+]?\d*\.?\d*$')
    numeric = pd.to_numeric(col.where(mask), errors="coerce")
    total = numeric.sum(skipna=True)
    print(total)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

